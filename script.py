# -*- coding: utf-8 -*-
"""
Created on Thu Aug 13 15:08:09 2020

@author: Raj kumar Nayak
Email: rajkumarnayak628@gmail.com
"""

from openpyxl import load_workbook
import xlrd
import tkinter
from tkinter import *
from tkinter import messagebox
import shutil
import os
from itertools import chain

workbook = load_workbook(filename=r"C:\Users\user\Desktop\Project\Updated CAN Matrix DBC Automation_Project_VB.xlsm")
worksheet=workbook.active
worksheet2=workbook[workbook.sheetnames[1]]
Filename=r"C:\Users\user\Desktop\Project\Ext\DBCFiles"
txt_file= open(Filename+".txt","w+")

new_symbols = [
    'NS_DESC_', 'CM_', 'BA_DEF_', 'BA_', 'VAL_', 'CAT_DEF_', 'CAT_',
    'FILTER', 'BA_DEF_DEF_', 'EV_DATA_', 'ENVVAR_DATA_', 'SGTYPE_',
    'SGTYPE_VAL_', 'BA_DEF_SGTYPE_', 'BA_SGTYPE_', 'SIG_TYPE_REF_',
    'VAL_TABLE_', 'SIG_GROUP_', 'SIG_VALTYPE_', 'SIGTYPE_VALTYPE_',
    'BO_TX_BU_', 'BA_DEF_REL_', 'BA_REL_', 'BA_DEF_DEF_REL_',
    'BU_SG_REL_', 'BU_EV_REL_', 'BU_BO_REL_', 'SG_MUL_VAL_'
]
lines=['        '+'VERSION '+r'""']
lines.append('\n\n')
#adding symbol
lines.append('NS_ :')
for symbol in new_symbols:
    lines.append('        ' + symbol )

# ! bit_timming--Line change need to be taken care 
#lines.append('\n')
lines.append("\nBS_:\n")
line=['BU_: ']  #created New List


#Add Nodes
Nod_col=int(worksheet.cell(1,5).value.split('=')[1])  #Declaring the Node's starting column 
ECU_count=int(worksheet.cell(2,5).value.split('=')[1]) #Getting the No of Nodes from Matrix
Node_col=Nod_col+ECU_count
for i in range(Node_col,Nod_col,-1):
    Node = worksheet.cell(3,i).value
    Node=re.sub(r"\W+|:", "_", Node)
    line.append(str(Node).replace(' ', '_') + ' ')
    
    

for listitem in lines:
        txt_file.write('%s\n' % listitem)
for listitem in line:
        txt_file.write('%s' % listitem)




#Message and signals 

dlc=8
Multiple_tx_count=0
row_value=4            #Starting Row Value
def row_increament():  #Function to check the Empty Row and strike off content
    global row_value
    while(worksheet.cell(row_value,5).value==None or worksheet.cell(row_value,5).font.strike!=None):
        row_value=row_value+1
            
    return row_value

def sig_row_increament(): #Check the Row after Last Signal and Limit is 5 Rows
    global row_value
    cou=0
    bool=True
    while(worksheet.cell(row_value,5).value==None and bool==True):
        row_value=row_value+1
        cou=cou+1
        if cou>4:
            bool=False
                
    return row_value

def strike_content():    #Function to check the strike off content
    global row_value
    while worksheet.cell(row_value,5).font.strike!=None:
        row_value=row_value+1
    return row_value

row_increament()
Tx=[]
while(worksheet.cell(row_value,5).value!=None):
    if(worksheet.cell(row_value,2).value!=None):
        m_line=['\nBO_ ']
        if (worksheet.cell(1,3).value=='Extend'):
            Hextodec=int(worksheet.cell(row_value,2).value,16)+pow(2,31)
        else:
            Hextodec=int(worksheet.cell(row_value,2).value,16)
        m_line.append(Hextodec)
        Mes=worksheet.cell(row_value,5).value
        Mes=re.sub(r"\W+|:", "", Mes)
        
        #Warning Box if the Length of Message will be Beyond 33 character
        
        if (len(Mes)>33):
            window =Tk()           #For Warning window
            window.eval('tk::PlaceWindow %s center' % window.winfo_toplevel())
            window.withdraw()
            messagebox.showwarning('Message Length', 'Please Decrease the Message Character')
            window.quit()
        
        m_line.append(' '+str(Mes)+':'+' '+str(dlc))
        count=0
        bool=False
        for i in range(Node_col,6,-1):        #Try to make a loop considering the no of Nodes
            if (worksheet.cell(row_value,i).value=='s'):
                sender_Node=worksheet.cell(3,i).value
                sender_Node=re.sub(r"\W+|:", "_", sender_Node)
                m_line.append(' '+str(sender_Node)+'\n')
                break
        
        
        
        #For Multiple transmitter
        Tx_sig=[]
        Node_count=0            
        for i in range(Node_col,Nod_col,-1):
            if (worksheet.cell(row_value,i).value=='s'):
                Node_name=worksheet.cell(3,i).value
                Node_name=re.sub(r"\W+|:", "_", Node_name)
                Tx_sig.append(str(Node_name))
                Node_count=Node_count+1
        
        if (Node_count>1):                         #Will Print Only when it has multiple transmitter
            Tx.append('BO_TX_BU_ ')
            Tx.append(str(Hextodec))
            Tx.append(' : ')
            Tx.append(' '.join(Tx_sig).replace(' ',',')+';'+'\n')
            Tot_tx=[]
            Tot_tx=list(chain(Tot_tx,Tx))
            Multiple_tx_count=Multiple_tx_count+1
              
            
           
        for listitem in m_line:   #Printing Message Line
            txt_file.write('%s' % listitem)   
        row_value=row_value+1     #Check new Row Value in same loop
    
    #Check if the signal row is not Empty
    row_increament()
    
    #Check if signal row is not striked off
    #strike_content()
    #print(row_value)
    
    sig_name=worksheet.cell(row_value,5).value       #Signal Name
    sig_name=re.sub(r"\W+|:", "", sig_name)
    #Warning Box if the Length of Signal will be Beyond 33 character
    
    if (len(sig_name)>33):
        window=Tk()
        window.eval('tk::PlaceWindow %s center' % window.winfo_toplevel())
        window.withdraw()
        messagebox.showwarning('Signal Length', 'Please Decrease the Signal Character')
        window.quit()
    sig_line=['SG_ ']                            #starting of signal line  
    sig_bit_no= int(worksheet.cell(row_value,Node_col+3).value)      
    if (worksheet.cell(row_value,Node_col+5).value.upper()=='UNSIGNED'):   #Taking Signed and Unsigned Value
        sign_type='+'
    else:
        sign_type='-'
    #if (worksheet.cell(row_value,5).value[0:11]=='Multiplexor'):
        
    
    Factor=worksheet.cell(row_value,Node_col+11).value   #Factor Value
    if Factor==None:                        #Default Value
        Factor=1
    offset=worksheet.cell(row_value,Node_col+12).value   #Offset Value
    if offset==None:
        offset=0                            #Default Value
    
    sig_bit_length= worksheet.cell(row_value,Node_col+1).value  #Signal bit Length
    if sig_bit_length>64:
        window=Tk()
        window.eval('tk::PlaceWindow %s center' % window.winfo_toplevel())
        window.withdraw()
        messagebox.showwarning('Signal Bit Length Exceeded', 'Please Decrease the Signal Bit length to 64bit as DLC=8')
        window.quit()
        
    #Add comma
    #For Multiplexor Signal
    if (worksheet.cell(row_value,5).value[0:11]=='Multiplexor'): 
        bool=True
    mul_count=''
    if bool==True:
        mul_count='m'+'%d' %count
    if (mul_count=='m0'):
        mul_count='M'
    
    #Byte Order
    Byte_type=worksheet.cell(row_value,Node_col+4).value
    if (Byte_type=='Motorola'):
        byte_order=0
    else:
        byte_order=1
    
    if (sign_type == "+"):                  #Physical Range
        physical_range_min = 0
        physical_range_max = ((pow(2,sig_bit_length) - 1)*Factor) + offset
    else :
        physical_range_min = -int(((pow(2,sig_bit_length-1))*Factor) + offset)
        physical_range_max = ((pow(2,sig_bit_length-1)-1)*Factor) + offset
        
        
    if (worksheet.cell(row_value,19).value==None):     #Unit Value
        unit_val=''
    else:
        unit_val=worksheet.cell(row_value,Node_col+9).value
    rec_Node=[]

    for i in range(Node_col,Nod_col,-1):
        if (worksheet.cell(row_value,i).value=='r'):    #Reciever Value
            reciever=worksheet.cell(3,i).value
            reciever=re.sub(r"\W+|:", "_", reciever)
            rec_Node.append(str(reciever))
    rec= [x.replace(' ', ',') for x in rec_Node]    #Need to add comma
    if len(rec_Node)==0:
        rec.append('VECTOR_XXX')               #IF their is no reciever Node
        
    
    sig_line.append(str(sig_name)+' '+str(mul_count)+':'+' '+str(sig_bit_no)+'|'+str(worksheet.cell(row_value,11).value)+'@'+str(byte_order)+sign_type+' '+'('+str(Factor)+','+str(offset)+')'+' '+'['+str(physical_range_min)+'|'+str(physical_range_max)+']'+' "'+str(unit_val)+'"  '+' '.join(rec).replace(' ',',') +'\n')
    
    row_value=row_value+1 
    sig_row_increament()    #Function to check Empty Rows, limit is 5
    count=count+1
    for listitem in sig_line:
        txt_file.write('%s' % listitem)
        
txt_file.write('\n')

#adding Tx Node
if Multiple_tx_count>0:
    for listitem in Tot_tx:
        txt_file.write('%s' %listitem)
txt_file.write('\n')   
#Adding Attributes
if (worksheet.cell(1,3).value=='Standard'):   #Checking the formate of Database
    txt_file.write('%s\n' %worksheet2.cell(2,8).value)
else:                                            #We Can add J1939 standard to by using elif
    txt_file.write('%s\n' %worksheet2.cell(3,8).value)
    
#Adding File name
attr=['BA_ "DBName" "']
attr.append(str(worksheet.cell(1,4).value)+'";'+'\n')  #Getting the File name from Excel sheet
for listitem in attr:
    txt_file.write('%s' %listitem)      

#CycleTime attribute
row_value=4

row_increament()
while(worksheet.cell(row_value,5).value!=None):
    if(worksheet.cell(row_value,2).value!=None):
        Hextodec=int(worksheet.cell(row_value,2).value,16)+pow(2,31)
        cycle_time=worksheet.cell(row_value,3).value       #Cycle time
          
        
        if (cycle_time!='Event'):
            cycle_count=0
            if cycle_time>50000: 
                window =Tk()           #For Warning window
                window.eval('tk::PlaceWindow %s center' % window.winfo_toplevel())
                window.withdraw()
                messagebox.showwarning('Cycle Time', 'Cycle Time Cannot be greater than 50000 Please change it')
                window.quit() 
            cyc_desc=['BA_ "GenMsgCycleTime" BO_ ']
            cyc_desc.append(str(Hextodec)+' '+str(cycle_time)+';'+'\n')
            for itemlist in cyc_desc:       #CycleTime Attribute
                txt_file.write('%s' %itemlist)   
        
        row_value=row_value+1
    row_value=row_value+1   
    sig_row_increament()
    
txt_file.write('\n') 
#Initial Value atrribute
row_value=4
cou=0
row_increament()
while(worksheet.cell(row_value,5).value!=None):
    if(worksheet.cell(row_value,2).value!=None):
        if (worksheet.cell(1,3).value=='Extend'):
            Hextodec=int(worksheet.cell(row_value,2).value,16)+pow(2,31)
        else:
            Hextodec=int(worksheet.cell(row_value,2).value,16)
        
        row_value = row_value+1
    row_increament()
    init_value=worksheet.cell(row_value,Node_col+13).value
    if (init_value!=0 and init_value!=None):    #Only print the value other than Zero
        sig_name=worksheet.cell(row_value,5).value
        sig_name=re.sub(r"\W+|:", "", sig_name)
        initial_attr= ['BA_ "GenSigStartValue" SG_ ']
        initial_attr.append(str(Hextodec)+' '+str(sig_name)+' '+str(init_value)+';'+'\n')
        for itemlist in initial_attr:
            txt_file.write( '%s' %itemlist)
    row_value = row_value+1
    sig_row_increament()
txt_file.write('\n')    
#Value Table   
row_value=4

row_increament()
while(worksheet.cell(row_value,5).value!=None):
    if(worksheet.cell(row_value,2).value!=None):
        if (worksheet.cell(1,3).value=='Extend'):
            Hextodec=int(worksheet.cell(row_value,2).value,16)+pow(2,31)
        else:
            Hextodec=int(worksheet.cell(row_value,2).value,16)
        row_value=row_value+1
    row_increament()
    sig_name=worksheet.cell(row_value,5).value
    sig_name=re.sub(r"\W+|:", "", sig_name)
    val=worksheet.cell(row_value,Node_col+6).value    #IS SPACE ALLOWED?
    if (val!=None and worksheet.cell(row_value,Node_col+6).font.strike==None ): #Checking the Blank and striked off cell
        #val.replace(' ','_')
        Val_tab=['VAL_ ']
        #comp=[]
        Val_tab.append(Hextodec)
        Val_tab.append(' '+''.join(str(sig_name)))
        val_desc=[]
        for i in range(0,val.count('\n')+1):
            first=val.split('\n')[i]            #GETTING THE NEW LINE
            hexa=first.split('=')[0]            #TAKING THE VALUE BEFOR EQUAL
            hexatodec=int(hexa,16)               #CONVERTING HEXA TO DEC
            disc=first.split('=')[1]            #TAKING THE DESCRIPTION VALUE AFTER EQUAL 
            
            val_desc.append(' '+str(hexatodec)+' '+'"'+str(disc)+'"')
        val_desc.reverse()   
        Val_tab.append(' '.join(val_desc)+';'+'\n')
        
        
        for itemlist in Val_tab:
            txt_file.write('%s' %itemlist)
    
    
    row_value=row_value+1
    sig_row_increament()


